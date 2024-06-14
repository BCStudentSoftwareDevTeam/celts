from flask import g
from peewee import fn, JOIN, ModelSelect
import re as regex
from openpyxl import load_workbook
from collections import defaultdict
from typing import DefaultDict, List, Dict, Any, Union

from app.models import mainDB
from app.models import DoesNotExist
from app.models.course import Course
from app.models.user import User
from app.models.term import Term
from app.models.courseInstructor import CourseInstructor
from app.models.courseParticipant import CourseParticipant
from app.models.courseStatus import CourseStatus
from app.models.courseQuestion import CourseQuestion
from app.models.attachmentUpload import AttachmentUpload
from app.models.questionNote import QuestionNote
from app.models.note import Note
from app.logic.createLogs import createAdminLog
from app.logic.fileHandler import FileHandler
from app.logic.term import addPastTerm
from app.logic.courseNameAndNumber import nameNumCombo

def getSLProposalInfoForUser(user: User) -> Dict[int, Dict[str, Any]]:
    """
    Given a user, returns a nested dictionary containing each course
    associated with the user and its course data. Used to populate 
    Service-Learning proposal table.
    """
    courses: List[Course] = list(Course.select(Course, Term, User, CourseStatus)
                                       .join(CourseInstructor).switch()
                                       .join(Term).switch()
                                       .join(CourseStatus).switch()
                                       .join(User)
                                       .where((CourseInstructor.user==user) | (Course.createdBy==user))
                                       .order_by(Course.term.desc(), Course.status))

    courseDict: Dict[int, Dict[str, Any]] = {}
    for course in courses:
        courseInstructors: List[CourseInstructor] = list(CourseInstructor.select(CourseInstructor, User).join(User).where(CourseInstructor.course==course))
        faculty: List[str] = [f"{instructor.user.firstName} {instructor.user.lastName}" for instructor in courseInstructors]


        courseDict[course.id] = {"id":course.id,
                                 "creator":f"{course.createdBy.firstName} {course.createdBy.lastName}",
                                 "coursename": nameNumCombo(course.courseName, course.courseAbbreviation),                               
                                 "faculty": faculty,
                                 "term": course.term,
                                 "status": course.status.status}
    return courseDict

def saveCourseParticipantsToDatabase(cpPreview: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]]) -> None:
    for term, terminfo in cpPreview.items():
        termObj: Term = Term.get_or_none(description = term) or addPastTerm(term)
        if not termObj:
            print(f"Unable to find or create term {term}")
            continue

        for course, courseInfo in terminfo['courses'].items():
            if 'errorMsg' in courseInfo and courseInfo['errorMsg']:
                print(f"Unable to save course {course}. {courseInfo['errorMsg']}")
                continue

            courseObj: Course = Course.get_or_create(
                 courseAbbreviation = course,
                 term = termObj, 
                 defaults = {"CourseName" : "",
                             "sectionDesignation" : "",
                             "courseCredit" : "1",
                             "term" : termObj,
                             "status" : 4,
                             "createdBy" : g.current_user,
                             "serviceLearningDesignatedSections" : "",
                             "previouslyApprovedDescription" : "" })[0]

            for userDict in courseInfo['students']:
                if userDict['errorMsg']:
                    print(f"Unable to save student. {userDict['errorMsg']}")
                    continue

                CourseParticipant.get_or_create(user=userDict['user'], 
                                                course=courseObj,
                                                hoursEarned=20)
                
def unapprovedCourses(termId: int) -> List[Course]:
    """
    Queries the database to get all the neccessary information for
    submitted/unapproved courses.
    """

    unapprovedCourses: List[Course] = list(Course.select(Course, Term, CourseStatus, fn.GROUP_CONCAT(" " ,User.firstName, " ", User.lastName).alias('instructors'))
                                                 .join(CourseInstructor, JOIN.LEFT_OUTER)
                                                 .join(User, JOIN.LEFT_OUTER).switch(Course)
                                                 .join(CourseStatus).switch(Course)
                                                 .join(Term)
                                                 .where(Term.id == termId,
                                                        Course.status.in_([CourseStatus.SUBMITTED, CourseStatus.IN_PROGRESS]))
                                                 .group_by(Course, Term, CourseStatus)
                                                 .order_by(Course.status))

    return unapprovedCourses

def approvedCourses(termId: int) -> List[Course]:
    """
    Queries the database to get all the necessary information for
    approved courses.
    """
    approvedCourses: List[Course] = list(Course.select(Course, Term, CourseStatus, fn.GROUP_CONCAT(" " ,User.firstName, " ", User.lastName).alias('instructors'))
                                               .join(CourseInstructor, JOIN.LEFT_OUTER)
                                               .join(User, JOIN.LEFT_OUTER).switch(Course)
                                               .join(CourseStatus).switch(Course)
                                               .join(Term)
                                               .where(Term.id == termId, Course.status == CourseStatus.APPROVED)
                                               .group_by(Course, Term, CourseStatus))

    return approvedCourses

def getImportedCourses(termId: int) -> List[Course]:
    """
    Queries the database to get all the necessary information for
    imported courses.
    """
    importedCourses: List[Course] = list(Course.select(Course, Term, CourseStatus, fn.GROUP_CONCAT(" " ,User.firstName, " ", User.lastName).alias('instructors'))
                                               .join(CourseInstructor, JOIN.LEFT_OUTER)
                                               .join(User, JOIN.LEFT_OUTER).switch(Course)
                                               .join(CourseStatus).switch(Course)
                                               .join(Term)
                                               .where(Term.id == termId, Course.status == CourseStatus.IMPORTED)
                                               .group_by(Course, Term, CourseStatus))

    return importedCourses

def getInstructorCourses() -> Dict[User, str]:
    """
    This function queries all of the course instructors and their classes and maps
    each instructor to its respective courses.
    """
    instructors: List[CourseInstructor] = list(CourseInstructor.select(CourseInstructor, User, Course)
                                                               .join(User).switch()
                                                               .join(Course))
    instructorToCoursesMap: DefaultDict[User, str] = defaultdict(list)

    for instructor in instructors:
        if instructor.course.courseName not in instructorToCoursesMap[instructor.user]:
            instructorToCoursesMap[instructor.user].append(instructor.course.courseName)

    return dict(instructorToCoursesMap)

########### Course Actions ###########

def renewProposal(courseID, term) -> Course:
    """
    Renews proposal of ID passed in for the selected term.
    Sets status to in progress.
    """
    oldCourse: Course = Course.get_by_id(courseID)
    newCourse: Course = Course.get_by_id(courseID)
    newCourse.id = None
    newCourse.term = Term.get_by_id(term)
    newCourse.status = CourseStatus.IN_PROGRESS
    newCourse.isPreviouslyApproved = True
    newCourse.save()
    questions: List[CourseQuestion] = list(CourseQuestion.select().where(CourseQuestion.course==oldCourse))
    for question in questions:
        CourseQuestion.create(course=newCourse.id,
                              questionContent=question.questionContent,
                              questionNumber=question.questionNumber)

    instructors = CourseInstructor.select().where(CourseInstructor.course==oldCourse.id)
    for instructor in instructors:
        CourseInstructor.create(course=newCourse.id,
                                user=instructor.user)

    return newCourse

def withdrawProposal(courseID) -> None:
    """
    Withdraws proposal of ID passed in. Removes foreign keys first.
    Key Dependencies: QuestionNote, CourseQuestion, CourseParticipant,
    CourseInstructor, Note
    """

    # delete syllabus
    try:
        syllabi: List[AttachmentUpload] = list(AttachmentUpload.select().where(AttachmentUpload.course==courseID))
        for syllabus in syllabi:
            FileHandler(courseId = courseID).deleteFile(syllabus.id)

    except DoesNotExist:
        print(f"File, {AttachmentUpload.fileName}, does not exist.")

    # delete course object
    course: Course = Course.get(Course.id == courseID)
    courseName: str = course.courseName
    questions: List[CourseQuestion] = CourseQuestion.select().where(CourseQuestion.course == course)
    notes: List[Note] = list(Note.select(Note.id)
                     .join(QuestionNote)
                     .where(QuestionNote.question.in_(questions))
                     .distinct())
    course.delete_instance(recursive=True)
    for note in notes:
        note.delete_instance()

    createAdminLog(f"Withdrew SLC proposal: {courseName}")

def createCourse(creator: str="No user provided") -> Course:
    """
    Creates and returns an empty, in progress course.
    """
    course: Course = Course.create(status=CourseStatus.IN_PROGRESS, createdBy=creator)
    for number in range(1, 7):
        CourseQuestion.create(course=course, questionNumber=number)

    return course

def updateCourse(courseData, attachments=None) -> Union[Course, bool]:
    """
    This function will take in courseData for the SLC proposal page and a dictionary
    of instuctors assigned to the course and update the information in the db.
    """
    with mainDB.atomic() as transaction:
        try:
            course: Course = Course.get_by_id(courseData['courseID'])

            for toggler in ["slSectionsToggle", "permanentDesignation"]:
                courseData.setdefault(toggler, "off")

            (Course.update(courseName=courseData["courseName"],
                           courseAbbreviation=courseData["courseAbbreviation"],
                           sectionDesignation=courseData["sectionDesignation"],
                           courseCredit=courseData["credit"],
                           isRegularlyOccurring=int(courseData["isRegularlyOccurring"]),
                           term=courseData['term'],
                           status=CourseStatus.SUBMITTED,
                           isPreviouslyApproved=int(courseData["isPreviouslyApproved"]),
                           previouslyApprovedDescription = courseData["previouslyApprovedDescription"],
                           isAllSectionsServiceLearning=("on" in courseData["slSectionsToggle"]),
                           serviceLearningDesignatedSections=courseData["slDesignation"],
                           isPermanentlyDesignated=("on" in courseData["permanentDesignation"]),
                           hasSlcComponent = int(courseData["hasSlcComponent"]))
                   .where(Course.id == course.id).execute())
            
            # update the existing entry with the new question responses
            for questionIndex in range(1, 7):
                (CourseQuestion.update(questionContent=courseData[f"{questionIndex}"])
                               .where((CourseQuestion.questionNumber == questionIndex) &
                                      (CourseQuestion.course==course)).execute())
                
            # delete all course instructors and create entries for the updated instructors 
            CourseInstructor.delete().where(CourseInstructor.course == course).execute()
            instructorList: List[str] = courseData.getlist('instructor[]')
            for instructor in instructorList:
                CourseInstructor.create(course=course, user=instructor)

            # save attachments to course if applicable
            if attachments:
                addFile: FileHandler = FileHandler(attachments, courseId=course.id)
                addFile.saveFiles()

            createAdminLog(f"Saved SLC proposal: {courseData['courseName']}")

            return Course.get_by_id(course.id)
        
        except Exception as e:
            print(e)
            transaction.rollback()
            return False
        
def editImportedCourses(courseData):
    """
        This function will take in courseData for the SLC proposal page and a dictionary
        of instructors assigned to the imported course after that one is edited 
        and update the information in the db.
    """

    with mainDB.atomic() as transaction:
        try:
            course = Course.get_by_id(courseData["courseId"])
            
            Course.update(courseName=courseData["courseName"]).where(Course.id == course.id).execute()

            (CourseParticipant.update(hoursEarned=courseData["hoursEarned"])
                              .where(CourseParticipant.course_id == course.id).execute())
            
            instructorList = []
            CourseInstructor.delete().where(CourseInstructor.course == course).execute() 
            
            if 'instructor[]' in courseData:
                instructorList = courseData.getlist('instructor[]') 
                
                for instructor in instructorList: 
                    # Checks that empty string is not added as a course instructor because some keys in the dictionary are empty string.
                    if instructor: 
                        CourseInstructor.create(course=course, user=instructor)         

            return Course.get_by_id(course.id)

        except Exception as e:
            print(e)
            transaction.rollback()
            return False
        
########### Course Actions ###########

def parseUploadedFile(filePath):
    """
    Parse an Excel document at the given `filePath` for courses and
    course participants.

    The return value will be a tuple. The second value is a list of 
    error message tuples. The first tuple value is the error message, 
    and the second is a 0 or 1 indicating whether the error is a 
    'general' error - 1 - or an error on a specific course, term, or 
    person.  The first value is a dictionary keyed by the term 
    description. Each value is another dictionary, with a key for 
    'displayMsg' and 'errorMsg', and a 'courses' key whose value is 
    a dictionary with keys for the courses in the term. Each course 
    has a 'displayMsg' and 'errorMsg' key, and a 'students' key 
    that has a list of dictionaries with 'user', 'displayMsg', and 
    'errorMsg' keys.
    E.g.,
    {
        "Fall 2021": {
            "displayMsg": "",
            "errorMsg": "",
            "courses": {
                "CSC 330": {
                    "displayMsg: "CSC 330 will be created",
                    "errorMsg: "",
                    "students": [
                        {'user':'ramsayb2', 
                         'displayMsg': 'Brian Ramsay',
                         'errorMsg': ''},
                        {'user':'B0073235', 
                         'displayMsg': '',
                         'errorMsg': 'ERROR: B0073235 does not exist!'}]
                },
                "CSC 226": {
                    "displayMsg": "CSC 226 matched to existing course 'Data Structures'.",
                    "errorMsg": "",
                    "students": [
                        {'user':'ramsayb2', 
                         'displayMsg': 'Brian Ramsay',
                         'errorMsg': ''},
                        {'user':'lamichhanes', 
                         'displayMsg': 'Sandesh Lamichhane',
                         'errorMsg': ''}]
                }
           }
        }
    }
    """
    excelData = load_workbook(filename=filePath)
    excelSheet = excelData.active

    result = {}
    errors = []
    term = ''
    course = ''
    cellRow = 0

    for row in excelSheet.iter_rows():
        cellRow += 1
        cellVal = row[0].value
        if not cellVal:
            continue

        # Look for a Term. Examples: Fall 2020 or Spring B 2021
        if regex.search(r"\b[a-zA-Z]{3,}( [AB])? \d{4}\b", str(cellVal)):
            errorMsg = ''
            if "Spring A" in cellVal or "Spring B" in cellVal:
                cellVal = "Spring " + cellVal.split()[-1]
            if "Fall A" in cellVal or "Fall B" in cellVal:
                cellVal = "Fall " + cellVal.split()[-1]
                
            if cellVal.split()[0] not in ["Summer", "Spring", "Fall", "May"]:
                errorMsg = f"ERROR: '{cellVal}' is not a valid term in row {cellRow}."
            else:
                latestTerm = Term.select().order_by(Term.termOrder.desc()).get()
                isFutureTerm = latestTerm.termOrder < Term.convertDescriptionToTermOrder(cellVal)
                if isFutureTerm:
                    errors.append(f"ERROR: '{cellVal}' is a future term in row {cellRow}.")
                else:
                    validTerm = Term.get_or_none(Term.description == cellVal)

            term = cellVal
            result[term] = {
                'displayMsg': term,
                'errorMsg': errorMsg,
                'courses': {}
            }
            if errorMsg:
                errors.append((errorMsg,0))

        # Look for a Course. Examples: FRN134 CSC 226
        elif regex.search(r"\b[A-Z]{2,4} ?\d{3}\b", str(cellVal)):
            errorMsg = displayMsg = ''
            if not term:
                displayMsg = cellVal
                errorMsg = "ERROR: No term was given for this course"
            else:
                existingCourse = Course.get_or_none(Course.courseAbbreviation == cellVal)
                displayMsg = f'{cellVal} will be created.'
                if existingCourse:
                    displayMsg = f"{cellVal} matched to the existing course {existingCourse.courseName}."

            course = cellVal
            result[term]['courses'][course] = {
                'displayMsg': displayMsg,
                'errorMsg': errorMsg,
                'students': []
            }
            if errorMsg:
                errors.append((errorMsg,0))

        # Look for a B-Number. Example: B00123456
        elif regex.search(r"\b[B]\d{8}\b", str(cellVal)):      
            errorMsg = displayMsg = ''
            if not course:
                errorMsg = "ERROR: No course is connected to this student"
            else:
                existingUser = User.get_or_none(User.bnumber == cellVal)
                if existingUser:
                    displayMsg = f"{existingUser.firstName} {existingUser.lastName}"
                    existingUser = existingUser.username
                else:             
                    errorMsg = f'ERROR: {row[1].value} with B# "{row[0].value}" does not exist.'

            result[term]['courses'][course]['students'].append({
                'user': (existingUser or cellVal),
                'displayMsg': displayMsg,
                'errorMsg': errorMsg})
            if errorMsg:
                errors.append((errorMsg,0))
            
        elif cellVal: # but didn't match the regex
            errors.append((f'ERROR: "{cellVal}" in row {cellRow} of the Excel document does not appear to be a term, course, or valid B#.',1))
        

    return result, errors