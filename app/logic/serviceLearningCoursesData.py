from flask import session, g
import re as regex
from openpyxl import load_workbook
from app.models.course import Course
from app.models.user import User
from app.models.term import Term
from app.models.courseInstructor import CourseInstructor
from app.models.courseParticipant import CourseParticipant
from app.models.courseStatus import CourseStatus
from app.models.courseQuestion import CourseQuestion
from app.models.questionNote import QuestionNote
from app.models.note import Note
from app.models.attachmentUpload import AttachmentUpload
from app.models.term import Term
from app.models import DoesNotExist
from app.logic.createLogs import createAdminLog
from app.logic.fileHandler import FileHandler
from app.logic.term import addPastTerm

def getServiceLearningCoursesData(user):
    """Returns dictionary with data used to populate Service-Learning proposal table"""
    courses = (Course.select(Course, Term, User, CourseStatus)
                     .join(CourseInstructor).switch()
                     .join(Term).switch()
                     .join(CourseStatus).switch()
                     .join(User)
                     .where((CourseInstructor.user==user)|(Course.createdBy==user))
                     .order_by(Course.term.desc(), Course.status))

    courseDict = {}
    for course in courses:
        otherInstructors = (CourseInstructor.select(CourseInstructor, User).join(User).where(CourseInstructor.course==course))
        faculty = [f"{instructor.user.firstName} {instructor.user.lastName}" for instructor in otherInstructors]


        courseDict[course.id] = {"id":course.id,
                                 "creator":f"{course.createdBy.firstName} {course.createdBy.lastName}",
                                 "name":course.courseName,
                                 "faculty": faculty,
                                 "term": course.term,
                                 "status": course.status.status}
    return courseDict

def withdrawProposal(courseID):
    """Withdraws proposal of ID passed in. Removes foreign keys first.
    Key Dependencies: QuestionNote, CourseQuestion, CourseParticipant,
    CourseInstructor, Note"""

    # delete syllabus
    try:
        syllabi = AttachmentUpload.select().where(AttachmentUpload.course==courseID)
        for syllabus in syllabi:
            FileHandler(courseId = courseID).deleteFile(syllabus.id)

    except DoesNotExist:
        print(f"File, {AttachmentUpload.fileName}, does not exist.")

    # delete course object
    course = Course.get(Course.id == courseID)
    courseName = course.courseName
    questions = CourseQuestion.select().where(CourseQuestion.course == course)
    notes = list(Note.select(Note.id)
                     .join(QuestionNote)
                     .where(QuestionNote.question.in_(questions))
                     .distinct())
    course.delete_instance(recursive=True)
    for note in notes:
        note.delete_instance()

    createAdminLog(f"Withdrew SLC proposal: {courseName}")

def renewProposal(courseID, term):
    """
    Renews proposal of ID passed in for the selected term.
    Sets status to in progress.
    """
    oldCourse = Course.get_by_id(courseID)
    newCourse = Course.get_by_id(courseID)
    newCourse.id = None
    newCourse.term = Term.get_by_id(term)
    newCourse.status = CourseStatus.IN_PROGRESS
    newCourse.isPreviouslyApproved = True
    newCourse.save()
    questions = CourseQuestion.select().where(CourseQuestion.course==oldCourse)
    for question in questions:
        CourseQuestion.create(course=newCourse.id,
                              questionContent=question.questionContent,
                              questionNumber=question.questionNumber)

    instructors = CourseInstructor.select().where(CourseInstructor.course==oldCourse.id)
    for instructor in instructors:
        CourseInstructor.create(course=newCourse.id,
                                user=instructor.user)

    return newCourse

def parseUploadedFile(filePath):
    """
    Parse an Excel document at the given `filePath` for courses and
    course participants.

    The return value will be a tuple. The second value is a list of 
    error message tuples. The first tuple value is the error message, 
    and the second is a 0 or 1 indicating whether the error is a 
    'general' error - 1 - or an error on a specific course, term, or 
    person.  The first value is a dictionary keyed by the term 
    description. Each value is another dictionary, with a key for 
    'displayMsg' and 'errorMsg', and a 'courses' key whose value is 
    a dictionary with keys for the courses in the term. Each course 
    has a 'displayMsg' and 'errorMsg' key, and a 'students' key 
    that has a list of dictionaries with 'user', 'displayMsg', and 
    'errorMsg' keys.
    E.g.,
    {
        "Fall 2021": {
            "displayMsg": "",
            "errorMsg": "",
            "courses": {
                "CSC 330": {
                    "displayMsg: "CSC 330 will be created",
                    "errorMsg: "",
                    "students": [
                        {'user':'ramsayb2', 
                         'displayMsg': 'Brian Ramsay',
                         'errorMsg': ''},
                        {'user':'B0073235', 
                         'displayMsg': '',
                         'errorMsg': 'ERROR: B0073235 does not exist!'}]
                },
                "CSC 226": {
                    "displayMsg": "CSC 226 matched to existing course 'Data Structures'.",
                    "errorMsg": "",
                    "students": [
                        {'user':'ramsayb2', 
                         'displayMsg': 'Brian Ramsay',
                         'errorMsg': ''},
                        {'user':'lamichhanes', 
                         'displayMsg': 'Sandesh Lamichhane',
                         'errorMsg': ''}]
                }
           }
        }
    }
    """
    excelData = load_workbook(filename=filePath)
    excelSheet = excelData.active

    result= {}
    errors = []
    term = ''
    course = ''
    cellRow = 0

    for row in excelSheet.iter_rows():
        cellRow += 1
        cellVal = row[0].value
        if not cellVal:
            continue

        # Look for a Term. Examples: Fall 2020 or Spring B 2021
        if regex.search(r"\b[a-zA-Z]{3,}( [AB])? \d{4}\b", str(cellVal)):
            errorMsg = ''
            if "Spring A" in cellVal or "Spring B" in cellVal:
                cellVal = "Spring " + cellVal.split()[-1]
            if "Fall A" in cellVal or "Fall B" in cellVal:
                cellVal = "Fall " + cellVal.split()[-1]
                
            if cellVal.split()[0] not in ["Summer", "Spring", "Fall", "May"]:
                errorMsg = f"ERROR: '{cellVal}' is not a valid term in row {cellRow}."
            else:
                latestTerm = Term.select().order_by(Term.termOrder.desc()).get()
                isFutureTerm = latestTerm.termOrder < Term.convertDescriptionToTermOrder(cellVal)
                if isFutureTerm:
                    errors.append(f"ERROR: '{cellVal}' is a future term in row {cellRow}.")
                else:
                    validTerm = Term.get_or_none(Term.description == cellVal)

            term = cellVal
            result[term] = {
                'displayMsg': term,
                'errorMsg': errorMsg,
                'courses': {}
            }
            if errorMsg:
                errors.append((errorMsg,0))

        # Look for a Course. Examples: FRN134 CSC 226
        elif regex.search(r"\b[A-Z]{2,4} ?\d{3}\b", str(cellVal)):
            errorMsg = displayMsg = ''
            if not term:
                displayMsg = cellVal
                errorMsg = "ERROR: No term was given for this course"
            else:
                existingCourse = Course.get_or_none(Course.courseAbbreviation == cellVal)
                displayMsg = f'{cellVal} will be created.'
                if existingCourse:
                    displayMsg = f"{cellVal} matched to the existing course {existingCourse.courseName}."

            course = cellVal
            result[term]['courses'][course] = {
                'displayMsg': displayMsg,
                'errorMsg': errorMsg,
                'students': []
            }
            if errorMsg:
                errors.append((errorMsg,0))

        # Look for a B-Number. Example: B00123456
        elif regex.search(r"\b[B]\d{8}\b", str(cellVal)):      
            errorMsg = displayMsg = ''
            if not course:
                errorMsg = "ERROR: No course is connected to this student"
            else:
                existingUser = User.get_or_none(User.bnumber == cellVal)
                if existingUser:
                    displayMsg = f"{existingUser.firstName} {existingUser.lastName}"
                    existingUser = existingUser.username
                else:             
                    errorMsg = f'ERROR: {row[1].value} with B# "{row[0].value}" does not exist.'

            result[term]['courses'][course]['students'].append({
                'user': (existingUser or cellVal),
                'displayMsg': displayMsg,
                'errorMsg': errorMsg})
            if errorMsg:
                errors.append((errorMsg,0))
            
        elif cellVal: # but didn't match the regex
            errors.append((f'ERROR: "{cellVal}" in row {cellRow} of the Excel document does not appear to be a term, course, or valid B#.',1))
        

    return result, errors

def saveCourseParticipantsToDatabase(cpPreview):
    for term,terminfo in cpPreview.items():
        termObj = Term.get_or_none(description = term) or addPastTerm(term)
        if not termObj:
            print(f"Unable to find or create term {term}")
            continue

        for course, courseinfo in terminfo['courses'].items():
            if 'errorMsg' in courseinfo and courseinfo['errorMsg']:
                print(f"Unable to save course {course}. {courseinfo['errorMsg']}")
                continue

            courseObj = Course.get_or_create(
                 courseAbbreviation = course,
                 term = termObj, 
                 defaults = {"CourseName" : "",
                             "sectionDesignation" : "",
                             "courseCredit" : "1",
                             "term" : termObj,
                             "status" : 4,
                             "createdBy" : g.current_user,
                             "serviceLearningDesignatedSections" : "",
                             "previouslyApprovedDescription" : "" })[0]

            for userDict in courseinfo['students']:
                if userDict['errorMsg']:
                    print(f"Unable to save student. {userDict['errorMsg']}")
                    continue

                CourseParticipant.get_or_create(user=userDict['user'], 
                                                course=courseObj,
                                                hoursEarned=20)
