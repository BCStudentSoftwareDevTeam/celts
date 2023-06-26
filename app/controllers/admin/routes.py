from flask import request, render_template, url_for, g, Flask, redirect
from flask import flash, abort, jsonify, session, send_file
from peewee import DoesNotExist, fn, IntegrityError
from playhouse.shortcuts import model_to_dict, dict_to_model
import json
from datetime import datetime, date
import os
import re

from app import app
from app.models.program import Program
from app.models.event import Event
from app.models.user import User
from app.models.eventTemplate import EventTemplate
from app.models.adminLogs import AdminLogs
from app.models.attachmentUpload import AttachmentUpload
from app.models.bonnerCohort import BonnerCohort
from app.models.certification import Certification
from app.models.user import User
from app.models.eventViews import EventView
from app.models.term import Term

from app.logic.userManagement import getAllowedPrograms, getAllowedTemplates
from app.logic.adminLogs import createLog
from app.logic.certification import getCertRequirements, updateCertRequirements
from app.logic.volunteers import getEventLengthInHours
from app.logic.utils import selectSurroundingTerms, getFilesFromRequest
from app.logic.events import deleteEvent, attemptSaveEvent, preprocessEventData, calculateRecurringEventFrequency, deleteEventAndAllFollowing, deleteAllRecurringEvents, getBonnerEvents,addEventView, getEventRsvpCountsForTerm
from app.logic.participants import getEventParticipants, getUserParticipatedTrainingEvents, checkUserRsvp, checkUserVolunteer
from app.logic.fileHandler import FileHandler
from app.logic.bonner import getBonnerCohorts, makeBonnerXls, rsvpForBonnerCohort
from app.controllers.admin import admin_bp
from openpyxl import load_workbook



@admin_bp.route('/switch_user', methods=['POST'])
def switchUser():
    if app.env == "production":
        print(f"An attempt was made to switch to another user by {g.current_user.username}!")
        abort(403)

    print(f"Switching user from {g.current_user} to",request.form['newuser'])
    session['current_user'] = model_to_dict(User.get_by_id(request.form['newuser']))

    return redirect(request.referrer)


@admin_bp.route('/eventTemplates')
def templateSelect():
    if g.current_user.isCeltsAdmin or g.current_user.isCeltsStudentStaff:
        allprograms = getAllowedPrograms(g.current_user)
        visibleTemplates = getAllowedTemplates(g.current_user)
        return render_template("/events/template_selector.html",
                                programs=allprograms,
                                templates=visibleTemplates)
    else:
        abort(403)


@admin_bp.route('/eventTemplates/<templateid>/create', methods=['GET','POST'])
@admin_bp.route('/eventTemplates/<templateid>/<programid>/create', methods=['GET','POST'])
def createEvent(templateid, programid=None):
    if not (g.current_user.isAdmin or g.current_user.isProgramManagerFor(programid)):
        abort(403)

    # Validate given URL
    program = None
    try:
        template = EventTemplate.get_by_id(templateid)
        if programid:
            program = Program.get_by_id(programid)
    except DoesNotExist as e:
        print("Invalid template or program id:", e)
        flash("There was an error with your selection. Please try again or contact Systems Support.", "danger")
        return redirect(url_for("admin.program_picker"))

    # Get the data from the form or from the template
    eventData = template.templateData

    if program:
        eventData["program"] = program

    if request.method == "GET":
        eventData['contactName'] = "CELTS Admin"
        eventData['contactEmail'] = app.config['celts_admin_contact']
        if program:
            eventData['location'] = program.defaultLocation
            if program.contactName:
                eventData['contactName'] = program.contactName
            if program.contactEmail:
                eventData['contactEmail'] = program.contactEmail

    # Try to save the form
    if request.method == "POST":
        eventData.update(request.form.copy())
        try:
            savedEvents, validationErrorMessage = attemptSaveEvent(eventData, getFilesFromRequest(request))

        except Exception as e:
            print("Error saving event:", e)
            savedEvents = False
            validationErrorMessage = "Unknown Error Saving Event. Please try again"

        if savedEvents:
            rsvpcohorts = request.form.getlist("cohorts[]")
            for year in rsvpcohorts:
                rsvpForBonnerCohort(int(year), savedEvents[0].id)

            noun = (eventData['isRecurring'] == 'on' and "Events" or "Event") # pluralize
            flash(f"{noun} successfully created!", 'success')

            if program:
                if len(savedEvents) > 1:
                    createLog(f"Created a recurring event, <a href=\"{url_for('admin.eventDisplay', eventId = savedEvents[0].id)}\">{savedEvents[0].name}</a>, for {program.programName}, with a start date of {datetime.strftime(eventData['startDate'], '%m/%d/%Y')}. The last event in the series will be on {datetime.strftime(savedEvents[-1].startDate, '%m/%d/%Y')}.")
                else:
                    createLog(f"Created <a href=\"{url_for('admin.eventDisplay', eventId = savedEvents[0].id)}\">{savedEvents[0].name}</a> for {program.programName}, with a start date of {datetime.strftime(eventData['startDate'], '%m/%d/%Y')}.")
            else:
                createLog(f"Created a non-program event, <a href=\"{url_for('admin.eventDisplay', eventId = savedEvents[0].id)}\">{savedEvents[0].name}</a>, with a start date of {datetime.strftime(eventData['startDate'], '%m/%d/%Y')}.")

            return redirect(url_for("admin.eventDisplay", eventId = savedEvents[0].id))
        else:
            flash(validationErrorMessage, 'warning')

    # make sure our data is the same regardless of GET or POST
    preprocessEventData(eventData)
    isProgramManager = g.current_user.isProgramManagerFor(programid)

    futureTerms = selectSurroundingTerms(g.current_term, prevTerms=0)

    requirements, bonnerCohorts = [], []
    if 'program' in eventData and eventData['program'].isBonnerScholars:
        requirements = getCertRequirements(Certification.BONNER)
        bonnerCohorts = getBonnerCohorts(limit=5)

    return render_template(f"/admin/{template.templateFile}",
            template = template,
            eventData = eventData,
            futureTerms = futureTerms,
            requirements = requirements,
            bonnerCohorts = bonnerCohorts,
            isProgramManager = isProgramManager)

@admin_bp.route('/event/<eventId>/view', methods=['GET'])
@admin_bp.route('/event/<eventId>/edit', methods=['GET','POST'])
def eventDisplay(eventId):
    pageViewsCount = EventView.select().where(EventView.event == eventId).count()
    if request.method == 'GET' and request.path == f'/event/{eventId}/view':
        viewer = g.current_user
        event = Event.get_by_id(eventId)
        addEventView(viewer,event) 
    # Validate given URL
    try:
        event = Event.get_by_id(eventId)
    except DoesNotExist as e:
        print(f"Unknown event: {eventId}")
        abort(404)

    notPermitted = not (g.current_user.isCeltsAdmin or g.current_user.isProgramManagerForEvent(event))
    if 'edit' in request.url_rule.rule and notPermitted:
        abort(403)

    eventData = model_to_dict(event, recurse=False)
    associatedAttachments = AttachmentUpload.select().where(AttachmentUpload.event == event)

    if request.method == "POST": # Attempt to save form
        eventData = request.form.copy()
        try:
            savedEvents, validationErrorMessage = attemptSaveEvent(eventData, getFilesFromRequest(request))

        except Exception as e:
            print("Error saving event:", e)
            savedEvents = False
            validationErrorMessage = "Unknown Error Saving Event. Please try again"


        if savedEvents:
            rsvpcohorts = request.form.getlist("cohorts[]")
            for year in rsvpcohorts:
                rsvpForBonnerCohort(int(year), event.id)

            flash("Event successfully updated!", "success")
            return redirect(url_for("admin.eventDisplay", eventId = event.id))
        else:
            flash(validationErrorMessage, 'warning')

    # make sure our data is the same regardless of GET and POST
    preprocessEventData(eventData)
    eventData['program'] = event.singleProgram
    futureTerms = selectSurroundingTerms(g.current_term)
    userHasRSVPed = checkUserRsvp(g.current_user, event)
    isPastEvent = event.isPast
    filepaths = FileHandler(eventId=event.id).retrievePath(associatedAttachments)
    isProgramManager = g.current_user.isProgramManagerFor(eventData['program'])

    requirements, bonnerCohorts = [], []
    if eventData['program'] and eventData['program'].isBonnerScholars:
        requirements = getCertRequirements(Certification.BONNER)
        bonnerCohorts = getBonnerCohorts(limit=5)
    
    rule = request.url_rule

    # Event Edit
    if 'edit' in rule.rule:
        return render_template("admin/createEvent.html",
                                eventData = eventData,
                                futureTerms=futureTerms,
                                isPastEvent = isPastEvent,
                                requirements = requirements,
                                bonnerCohorts = bonnerCohorts,
                                userHasRSVPed = userHasRSVPed,
                                isProgramManager = isProgramManager,
                                filepaths = filepaths)
    # Event View
    else:
        # get text representations of dates
        eventData['timeStart'] = event.timeStart.strftime("%-I:%M %p")
        eventData['timeEnd'] = event.timeEnd.strftime("%-I:%M %p")
        eventData["startDate"] = event.startDate.strftime("%m/%d/%Y")

        # Identify the next event in a recurring series
        if event.recurringId:
            eventSeriesList = list(Event.select().where(Event.recurringId == event.recurringId).order_by(Event.startDate))
            eventIndex = eventSeriesList.index(event)
            if len(eventSeriesList) != (eventIndex + 1):
                eventData["nextRecurringEvent"] = eventSeriesList[eventIndex + 1]

        currentEventRsvpAmount = getEventRsvpCountsForTerm(g.current_term)

        UserParticipatedTrainingEvents = getUserParticipatedTrainingEvents(eventData['program'], g.current_user, g.current_term)
        return render_template("eventView.html",
                                eventData = eventData,
                                isPastEvent = isPastEvent,
                                userHasRSVPed = userHasRSVPed,
                                programTrainings = UserParticipatedTrainingEvents,
                                currentEventRsvpAmount = currentEventRsvpAmount,
                                isProgramManager = isProgramManager,
                                filepaths = filepaths,
                                pageViewsCount= pageViewsCount)

@admin_bp.route('/event/<eventId>/delete', methods=['POST'])
def deleteRoute(eventId):
    try:
        deleteEvent(eventId)
        flash("Event successfully deleted.", "success")
        return redirect(url_for("main.events", selectedTerm=g.current_term))

    except Exception as e:
        print('Error while canceling event:', e)
        return "", 500
@admin_bp.route('/event/<eventId>/deleteEventAndAllFollowing', methods=['POST'])
def deleteEventAndAllFollowingRoute(eventId):
    try:
        deleteEventAndAllFollowing(eventId)
        flash("Events successfully deleted.", "success")
        return redirect(url_for("main.events", selectedTerm=g.current_term))

    except Exception as e:
        print('Error while canceling event:', e)
        return "", 500
@admin_bp.route('/event/<eventId>/deleteAllRecurring', methods=['POST'])
def deleteAllRecurringEventsRoute(eventId):
    try:
        deleteAllRecurringEvents(eventId)
        flash("Events successfully deleted.", "success")
        return redirect(url_for("main.events", selectedTerm=g.current_term))

    except Exception as e:
        print('Error while canceling event:', e)
        return "", 500

@admin_bp.route('/makeRecurringEvents', methods=['POST'])
def addRecurringEvents():
    recurringEvents = calculateRecurringEventFrequency(preprocessEventData(request.form.copy()))
    return json.dumps(recurringEvents, default=str)


@admin_bp.route('/userProfile', methods=['POST'])
def userProfile():
    volunteerName= request.form.copy()
    if volunteerName['searchStudentsInput']:
        username = volunteerName['searchStudentsInput'].strip("()")
        user=username.split('(')[-1]
        return redirect(url_for('main.viewUsersProfile', username=user))
    else:
        flash(f"Please enter the first name or the username of the student you would like to search for.", category='danger')
        return redirect(url_for('admin.studentSearchPage'))

@admin_bp.route('/search_student', methods=['GET'])
def studentSearchPage():
    if g.current_user.isAdmin:
        return render_template("/admin/searchStudentPage.html")
    abort(403)

@admin_bp.route('/addParticipants', methods = ['GET'])
def addParticipants():
    '''Renders the page, will be removed once merged with full page'''

    return render_template('addParticipants.html',
                            title="Add Participants")

@admin_bp.route('/adminLogs', methods = ['GET', 'POST'])
def adminLogs():
    if g.current_user.isCeltsAdmin:
        allLogs = AdminLogs.select(AdminLogs, User).join(User).order_by(AdminLogs.createdOn.desc())
        return render_template("/admin/adminLogs.html",
                                allLogs = allLogs)
    else:
        abort(403)

@admin_bp.route("/deleteEventFile", methods=["POST"])
def deleteEventFile():
    fileData= request.form
    eventfile=FileHandler(eventId=fileData["eventId"])
    eventfile.deleteFile(fileData["fileId"])
    return ""

@admin_bp.route("/uploadCourseParticipant", methods= ["POST"])
def addCourseFile():
    fileData = request.files['addCourseParticipant']
    filePath = os.path.join(app.config["files"]["base_path"], fileData.filename)
    fileData.save(filePath)
    excelData = load_workbook(filename=filePath)
    excelSheet = excelData.active

    courseAbrev = "" 
    term = ""


    row = ['course name', 'course number', 'faculty', 'term', 'previously']
    row[0]
    termReg = r"\b[a-zA-Z]{3,}\s\d{4}\b"
    courseReg = r"\b[A-Z]{2,4}\s\d{3}\b"
    bnumberReg = r"\b[B]\d{8}\b"
    for row in excelSheet.iter_rows():
        cellVal = row[0].value
    
        if re.search(termReg, str(cellVal)):
            # get term obj from database
            termObj = Term.select()
            term = termReg
          

            


        elif re.search(courseReg, str(cellVal)):
            # get course obj from database, create if doesn't exist yet
            course_abrev = courseReg
          
        
        elif re.search(bnumberReg, str(cellVal)):
            # get studentname from database
            # add term,course,student to courseparticipant
   

        else:
            print("/////////////////////// INVALID INPUT //////////////////////////\n")



    studentbnumber(course, term``)

    
    print(":::::::::::::::::::::::::::::::::::::")
    


    

    # if request.method == POST:
    #     file = request.files.get["addCourseParticipant"]
    #     if file: 
    #          file.save("uploads/" + file.filename)
    #          return "File uploaded successfully."
    os.remove(filePath)
    
    return redirect(url_for("main.getAllCourseInstructors"))



    

@admin_bp.route("/manageBonner")
def manageBonner():
    if not g.current_user.isCeltsAdmin:
        abort(403)

    return render_template("/admin/bonnerManagement.html",
                           cohorts=getBonnerCohorts(),
                           events=getBonnerEvents(g.current_term),
                           requirements = getCertRequirements(certification=Certification.BONNER))

@admin_bp.route("/bonner/<year>/<method>/<username>", methods=["POST"])
def updatecohort(year, method, username):
    if not g.current_user.isCeltsAdmin:
        abort(403)

    try:
        user = User.get_by_id(username)
    except:
        abort(500)

    if method == "add":
        try:
            BonnerCohort.create(year=year, user=user)
        except IntegrityError as e:
            # if they already exist, ignore the error
            pass
    elif method == "remove":
        BonnerCohort.delete().where(BonnerCohort.user == user, BonnerCohort.year == year).execute()
    else:
        abort(500)

    return ""

@admin_bp.route("/bonnerxls")
def bonnerxls():
    if not g.current_user.isCeltsAdmin:
        abort(403)

    newfile = makeBonnerXls()
    return send_file(open(newfile, 'rb'), download_name='BonnerStudents.xlsx', as_attachment=True)

@admin_bp.route("/saveRequirements/<certid>", methods=["POST"])
def saveRequirements(certid):
    if not g.current_user.isCeltsAdmin:
        abort(403)

    newRequirements = updateCertRequirements(certid, request.get_json())

    return jsonify([requirement.id for requirement in newRequirements])
